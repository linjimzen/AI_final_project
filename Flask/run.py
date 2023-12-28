#https://www.youtube.com/watch?v=FUj88Q_L62k
#postman測試
from flask import Flask, request, jsonify,render_template,redirect,url_for
from flask_cors import CORS
import numpy as np
import os
import time
import model,model2
import openpyxl
app = Flask(__name__)
CORS(app)

@app.route("/")
def index():
    return render_template('index.html')

@app.route('/upload', methods=['get'])
def upload_page():
    return render_template('upload.html')
@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files['filename']  # 獲取上傳的文件
    #file.save(os.path.join(file.filename))
    # 在這裡處理上傳的文件，例如保存到本地或進行其他處理
    wb = openpyxl.load_workbook(file)
    sheet = wb[wb.sheetnames[0]]
    maxrow=sheet.max_row
    maxcolumn=sheet.max_column
    print(maxcolumn,maxrow)
    # 把excel的text存到text_list當中
    text_list=[]
    
    start_time = time.time()
    for text in range(1,maxrow+1):
        text_list.append(str(sheet.cell(text,1).value))
    #  丟入model預測
    pred_list=[]
    for text in text_list:
        pred_list.append(model2.pred(text))

    cost_time = time.time()-start_time 
    #合併兩個list
    combine_list=zip(text_list,pred_list)
    num_neg=0
    num_pos=0
    for i in pred_list:
        if(i=="Negative"):
            num_neg+=1
        else:
            num_pos+=1
    return render_template('upload.html',combine_list=combine_list,num_neg=num_neg,num_pos=num_pos,time=format(cost_time, ".3f"))


if __name__ == '__main__':
    app.run(host='0.0.0.0',port=3000,debug=True)

